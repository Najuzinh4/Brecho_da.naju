import os
import pandas as pd
from openpyxl import load_workbook, Workbook
from datetime import datetime

EXCEL_PATH = "pecas.xlsx"
HEADERS = ["codigo", "descricao", "categoria", "fornecedora",
           "comissao_pct", "preco", "status", "vendido_em"]


def init_excel():
    if not os.path.exists(EXCEL_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = "Peças"
        ws.append(HEADERS)
        wb.save(EXCEL_PATH)


def load_pecas():
    init_excel()
    df = pd.read_excel(EXCEL_PATH, sheet_name="Peças", dtype={"codigo": str})
    df = df.fillna("")
    return df


def save_peca(row: dict):
    init_excel()
    wb = load_workbook(EXCEL_PATH)
    ws = wb["Peças"]
    ws.append([
        row["codigo"], row["descricao"], row["categoria"],
        row["fornecedora"], float(row["comissao_pct"]),
        float(row["preco"]), "disponivel", ""
    ])
    wb.save(EXCEL_PATH)


def update_status(codigo: str, novo_status: str):
    wb = load_workbook(EXCEL_PATH)
    ws = wb["Peças"]
    headers = [c.value for c in ws[1]]
    idx_codigo  = headers.index("codigo") + 1
    idx_status  = headers.index("status") + 1
    idx_vendido = headers.index("vendido_em") + 1
    for row in ws.iter_rows(min_row=2):
        if str(row[idx_codigo - 1].value) == str(codigo):
            row[idx_status - 1].value = novo_status
            if novo_status == "vendido":
                row[idx_vendido - 1].value = datetime.now().strftime("%Y-%m-%d %H:%M")
            else:
                row[idx_vendido - 1].value = ""
            break
    wb.save(EXCEL_PATH)


def delete_peca(codigo: str):
    wb = load_workbook(EXCEL_PATH)
    ws = wb["Peças"]
    headers = [c.value for c in ws[1]]
    idx_codigo = headers.index("codigo") + 1
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if str(row[idx_codigo - 1].value) == str(codigo):
            ws.delete_rows(i)
            break
    wb.save(EXCEL_PATH)
