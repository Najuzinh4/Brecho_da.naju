from flask import Flask, render_template, request, jsonify
import pandas as pd
import openpyxl
from openpyxl import load_workbook, Workbook
from datetime import datetime
import os
import json
import requests
from code_service import proximo_codigo_para

app = Flask(__name__)
EXCEL_PATH = "pecas.xlsx"

# ── helpers ──────────────────────────────────────────────────────────────────

def init_excel():
    if not os.path.exists(EXCEL_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = "Peças"
        ws.append(["codigo", "descricao", "categoria", "fornecedora",
                   "comissao_pct", "preco", "status", "vendido_em"])
        wb.save(EXCEL_PATH)

def load_pecas():
    init_excel()
    df = pd.read_excel(EXCEL_PATH, sheet_name="Peças", dtype={"codigo": str})
    df = df.fillna("")
    return df

def save_peca(row: dict):
    init_excel()
    wb = load_workbook(EXCEL_PATH)
    ws = wb["Peças"]
    ws.append([
        row["codigo"], row["descricao"], row["categoria"],
        row["fornecedora"], float(row["comissao_pct"]),
        float(row["preco"]), "disponivel", ""
    ])
    wb.save(EXCEL_PATH)

def update_status(codigo: str, novo_status: str):
    wb = load_workbook(EXCEL_PATH)
    ws = wb["Peças"]
    headers = [c.value for c in ws[1]]
    idx_codigo   = headers.index("codigo") + 1
    idx_status   = headers.index("status") + 1
    idx_vendido  = headers.index("vendido_em") + 1
    for row in ws.iter_rows(min_row=2):
        if str(row[idx_codigo - 1].value) == str(codigo):
            row[idx_status - 1].value = novo_status
            if novo_status == "vendido":
                row[idx_vendido - 1].value = datetime.now().strftime("%Y-%m-%d %H:%M")
            else:
                row[idx_vendido - 1].value = ""
            break
    wb.save(EXCEL_PATH)

def delete_peca(codigo: str):
    wb = load_workbook(EXCEL_PATH)
    ws = wb["Peças"]
    headers = [c.value for c in ws[1]]
    idx_codigo = headers.index("codigo") + 1
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if str(row[idx_codigo - 1].value) == str(codigo):
            ws.delete_rows(i)
            break
    wb.save(EXCEL_PATH)

def calcular_relatorio(df):
    vendidas = df[df["status"] == "vendido"].copy()
    vendidas["comissao_pct"] = pd.to_numeric(vendidas["comissao_pct"], errors="coerce").fillna(40)
    vendidas["preco"] = pd.to_numeric(vendidas["preco"], errors="coerce").fillna(0)
    vendidas["val_naju"]       = vendidas["preco"] * vendidas["comissao_pct"] / 100
    vendidas["val_fornecedora"] = vendidas["preco"] - vendidas["val_naju"]

    total_vendido  = vendidas["preco"].sum()
    total_naju     = vendidas["val_naju"].sum()

    por_fornecedora = (
        vendidas.groupby("fornecedora")
        .agg(pecas=("codigo", "count"),
             total_vendas=("preco", "sum"),
             total_a_pagar=("val_fornecedora", "sum"))
        .reset_index()
        .to_dict(orient="records")
    )

    return {
        "total_vendido": round(total_vendido, 2),
        "total_naju": round(total_naju, 2),
        "por_fornecedora": [
            {**r,
             "total_vendas": round(r["total_vendas"], 2),
             "total_a_pagar": round(r["total_a_pagar"], 2)}
            for r in por_fornecedora
        ]
    }

# ── Claude API ────────────────────────────────────────────────────────────────

def claude_ajuda(tipo: str, descricao: str, categoria: str = "", preco: str = ""):
    system = (
        "Você é assistente da Brechó da Najú, brechó brasileiro Y2K/grunge/upcycling. "
        "A marca fala em português, lowercase, tom íntimo e slangy. "
        "Usa ★ ✦ como pontuação. Nunca usa corporativo."
    )

    if tipo == "descricao_instagram":
        prompt = (
            f"Crie uma legenda curta pra Instagram pra essa peça do brechó:\n"
            f"Peça: {descricao}\nCategoria: {categoria}\nPreço: R$ {preco}\n\n"
            f"Máximo 5 linhas. Estilo Najú: lowercase, estrelas, hashtags no final."
        )
    elif tipo == "sugerir_preco":
        prompt = (
            f"Sugira um preço justo pra brechó brasileiro pra essa peça:\n"
            f"{descricao}\nCategoria: {categoria}\n\n"
            f"Responda só com o valor em reais (ex: 45.00) e uma linha explicando."
        )
    elif tipo == "categorizar":
        prompt = (
            f"Categorize essa peça de brechó em UMA dessas categorias: "
            f"topo, calça, vestido, saia, jaqueta, acessório, sapato, bolsa, outro.\n"
            f"Peça: {descricao}\n\nResponda só com a categoria, sem mais nada."
        )
    else:
        return {"erro": "tipo inválido"}

    try:
        resp = requests.post(
            "https://api.anthropic.com/v1/messages",
            headers={"Content-Type": "application/json"},
            json={
                "model": "claude-sonnet-4-20250514",
                "max_tokens": 300,
                "system": system,
                "messages": [{"role": "user", "content": prompt}]
            },
            timeout=15
        )
        data = resp.json()
        texto = data["content"][0]["text"]
        return {"resultado": texto}
    except Exception as e:
        return {"erro": str(e)}

# ── Rotas ─────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/api/pecas", methods=["GET"])
def get_pecas():
    df = load_pecas()
    return jsonify(df.to_dict(orient="records"))

@app.route("/api/pecas", methods=["POST"])
def post_peca():
    data = request.json
    # se não vier código, gera automaticamente
    if not data.get("codigo"):
        df = load_pecas()
        data["codigo"] = proximo_codigo_para(data.get("fornecedora", "X"), df)
    save_peca(data)
    return jsonify({"ok": True, "codigo": data["codigo"]})

@app.route("/api/codigo-preview", methods=["POST"])
def codigo_preview():
    """Retorna o próximo código para uma fornecedora sem salvar nada."""
    fornecedora = request.json.get("fornecedora", "").strip()
    if not fornecedora:
        return jsonify({"codigo": ""})
    df = load_pecas()
    codigo = proximo_codigo_para(fornecedora, df)
    return jsonify({"codigo": codigo})

@app.route("/api/pecas/<codigo>/vender", methods=["POST"])
def vender(codigo):
    update_status(codigo, "vendido")
    return jsonify({"ok": True})

@app.route("/api/pecas/<codigo>/devolver", methods=["POST"])
def devolver(codigo):
    update_status(codigo, "disponivel")
    return jsonify({"ok": True})

@app.route("/api/pecas/<codigo>", methods=["DELETE"])
def deletar(codigo):
    delete_peca(codigo)
    return jsonify({"ok": True})

@app.route("/api/relatorio", methods=["GET"])
def relatorio():
    df = load_pecas()
    return jsonify(calcular_relatorio(df))

@app.route("/api/claude", methods=["POST"])
def claude_route():
    data = request.json
    resultado = claude_ajuda(
        tipo=data.get("tipo"),
        descricao=data.get("descricao", ""),
        categoria=data.get("categoria", ""),
        preco=data.get("preco", "")
    )
    return jsonify(resultado)

if __name__ == "__main__":
    init_excel()
    print("\n★ Brechó da Najú PDV rodando em http://localhost:5000\n")
    app.run(debug=True, port=5000)
